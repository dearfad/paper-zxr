"""图像验证模块 - 重构版主入口

【模块概述】
本模块是细胞显微图像质量自动化验证系统的核心入口脚本。它整合了文件名规范校验、图像读取测试、完整性检查及多维度质量检测算法，
旨在对批量原始图像数据进行预处理前的质量筛查，确保后续分析流程的数据可靠性。

【核心功能】
1. 文件名合规性校验：基于正则表达式验证命名规范（细胞系、编号、时间、孔位等字段）。
2. 图像基础信息提取：读取并验证图像格式、尺寸、色彩模式及可读性。
3. 多维质量检测：
   - 模糊检测：基于拉普拉斯方差算法评估图像清晰度。
   - 瑕疵检测：识别并计数图像中的气泡和划痕缺陷。
   - 光学属性分析：计算亮度均值/标准差及对比度方差。
   - 覆盖率评估：计算有效细胞区域占比，识别低覆盖率样本。
4. 综合判定：根据上述指标自动判定图像是否“合格”。
5. 批量处理与报告生成：支持目录级批量扫描，采样分析，并将详细检测结果与统计汇总导出为 Excel 报表。

【工作流程】
1. 路径初始化：动态计算项目根目录并配置到 sys.path，解决模块化导入问题。
2. 单图分析 (analyze_single_image)：串联文件名校验、基础读取及各子检测器函数。
3. 批量执行 (validate_directory)：遍历目标目录，执行单图分析并实时打印进度。
4. 统计计算 (calculate_statistics)：聚合批次数据，计算合格率及各类缺陷分布。
5. 结果导出 (save_results_to_excel)：生成包含详细数据行和统计摘要表的 Excel 文件。

【模块依赖】
- 本地子模块：libs.preprocessing 下的各类检测器 (blur, bubble, scratch, coverage 等) 及工具函数。
- 第三方库：numpy (数值计算), openpyxl (Excel 生成)。

【使用示例】
# 直接运行脚本 (默认扫描配置目录)
python validate_images.py

# 代码调用
from preprocessing.validate_images import main, validate_directory
from pathlib import Path

# 指定自定义目录运行
main(target_dir=Path("./custom_data/images"))
"""

import random
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

# =============================================================================
# 路径配置与导入修复
# =============================================================================


import numpy as np
from openpyxl import Workbook
from libs.preprocessing.check_filename import check_filename, FILENAME_PATTERN
from libs.preprocessing.check_blur import check_blur, DEFAULT_BLUR_THRESHOLD
from libs.preprocessing.check_info import get_image_basic_info, load_image_bgr
from libs.preprocessing.check_bubble import detect_bubbles, DEFAULT_MIN_BUBBLE_AREA
from libs.preprocessing.check_scratch import (
    detect_scratches,
    DEFAULT_MIN_SCRATCH_LENGTH,
)
from libs.preprocessing.check_brightness import detect_brightness, detect_contrast
from libs.preprocessing.check_coverage import (
    detect_cell_coverage,
    DEFAULT_CELL_COVERAGE_THRESHOLD,
)
from libs.preprocessing.check_background import (
    detect_background_uniformity,
    DEFAULT_BACKGROUND_STD_THRESHOLD,
)
from libs.preprocessing.check_clump import (
    detect_cell_clumping,
    DEFAULT_CLUMP_SIZE_THRESHOLD,
)
from libs.preprocessing.check_defocus import (
    detect_partial_defocus,
    DEFAULT_DEFOCUS_RATIO_THRESHOLD,
)
from libs.preprocessing.check_media import (
    detect_media_color,
    DEFAULT_COLOR_DEVIATION_THRESHOLD,
)


# =============================================================================
# 配置常量 (路径相关)
# =============================================================================

# 默认原始图片目录 (相对于当前项目根目录调整)
# 修改：确保基于当前文件所在目录解析，避免运行环境不同导致的路径错误

CURRENT_FILE_DIR = Path(__file__).parent
DEFAULT_RAW_IMAGE_DIR = CURRENT_FILE_DIR.parent / "data" / "raw"
# 默认结果输出目录
DEFAULT_RESULT_DIR = CURRENT_FILE_DIR.parent / "data" / "results"


# =============================================================================
# 1. 单张图片完整分析 (协调者)
# =============================================================================


def analyze_single_image(
    image_path: Path,
    pattern: str = FILENAME_PATTERN,
    blur_threshold: float = DEFAULT_BLUR_THRESHOLD,
    min_bubble_area: int = DEFAULT_MIN_BUBBLE_AREA,
    min_scratch_length: int = DEFAULT_MIN_SCRATCH_LENGTH,
    coverage_threshold: float = DEFAULT_CELL_COVERAGE_THRESHOLD,
    # 新增参数：背景、聚集、离焦、颜色阈值
    bg_std_threshold: float = DEFAULT_BACKGROUND_STD_THRESHOLD,
    clump_size_threshold: int = DEFAULT_CLUMP_SIZE_THRESHOLD,
    defocus_ratio_threshold: float = DEFAULT_DEFOCUS_RATIO_THRESHOLD,
    color_deviation_threshold: float = DEFAULT_COLOR_DEVIATION_THRESHOLD,
) -> Dict[str, Any]:
    """对单张图片执行全流程分析"""
    result = {
        "filename": image_path.name,
        "file_path": str(image_path.absolute()),
    }

    # 1. 文件名验证
    name_info = check_filename(image_path.name, pattern)
    result["filename_valid"] = name_info.get("is_valid", False)
    fields = ["cell", "id", "time", "passage", "well", "location", "magnification"]
    for field in fields:
        result[field] = name_info.get(field, "") if result["filename_valid"] else ""

    # 2. 基础信息读取
    info = get_image_basic_info(image_path)
    result["readable"] = info["success"]
    result["error"] = info.get("error", "") if not info["success"] else ""

    if not info["success"]:
        result.update(
            {
                "image_format": "",
                "image_mode": "",
                "width": 0,
                "height": 0,
                "is_qualified": False,
                "is_blur": False,
                "variance": 0.0,
                "bubble_count": 0,
                "scratch_count": 0,
                "brightness_mean": 0.0,
                "brightness_std": 0.0,
                "contrast_variance": 0.0,
                "coverage_rate": 0.0,
                "is_low_coverage": True,
                # 新增字段初始化 (失败情况)
                "bg_uniformity_std": 0.0,
                "is_bg_non_uniform": True,
                "clump_count": 0,
                "has_clumps": False,
                "defocus_ratio": 0.0,
                "is_partial_defocus": True,
                "media_color_score": 0.0,
                "is_media_abnormal": True,
            }
        )
        return result

    result.update(
        {
            "image_format": info["format"],
            "image_mode": info["mode"],
            "width": info["width"],
            "height": info["height"],
        }
    )

    # 3. 质量检测 (调用子模块函数)
    try:
        img_bgr = check_blur(image_path)

        blur_res = check_blur(img_bgr, blur_threshold)
        result["is_blur"] = blur_res["is_blur"]
        result["variance"] = blur_res["variance"]

        bubble_res = detect_bubbles(img_bgr, min_bubble_area)
        result["bubble_count"] = bubble_res["count"]

        scratch_res = detect_scratches(img_bgr, min_scratch_length)
        result["scratch_count"] = scratch_res["count"]

        bright_res = detect_brightness(img_bgr)
        result["brightness_mean"] = bright_res["mean"]
        result["brightness_std"] = bright_res["std"]

        contrast_res = detect_contrast(img_bgr)
        result["contrast_variance"] = contrast_res["variance"]

        coverage_res = detect_cell_coverage(img_bgr, coverage_threshold)
        result["coverage_rate"] = coverage_res["coverage_rate"]
        result["is_low_coverage"] = coverage_res["is_low_coverage"]

        # --- 新增检测逻辑 ---

        # 背景均匀性检测
        bg_res = detect_background_uniformity(img_bgr, bg_std_threshold)
        result["bg_uniformity_std"] = bg_res["std_dev"]
        result["is_bg_non_uniform"] = bg_res["is_non_uniform"]

        # 细胞聚集/团块检测
        clump_res = detect_cell_clumping(img_bgr, clump_size_threshold)
        result["clump_count"] = clump_res["count"]
        result["has_clumps"] = clump_res["has_clumps"]

        # 多焦点/局部离焦检测
        defocus_res = detect_partial_defocus(img_bgr, defocus_ratio_threshold)
        result["defocus_ratio"] = defocus_res["ratio"]
        result["is_partial_defocus"] = defocus_res["is_defocused"]

        # 培养基状态检测 (颜色分析)
        media_res = detect_media_color(img_bgr, color_deviation_threshold)
        result["media_color_score"] = media_res["score"]
        result["is_media_abnormal"] = media_res["is_abnormal"]
        # ------------------

        # 综合判定：更新合格标准，纳入新检测项
        result["is_qualified"] = (
            not result["is_blur"]
            and result["bubble_count"] == 0
            and result["scratch_count"] == 0
            and not result["is_low_coverage"]
            and not result["is_bg_non_uniform"]
            and not result["has_clumps"]
            and not result["is_partial_defocus"]
            and not result["is_media_abnormal"]
        )

    except Exception as e:
        result["is_qualified"] = False
        result["error"] = f"质量检测异常：{str(e)}" + (
            f" (原错误：{info.get('error', '')})" if not info["success"] else ""
        )
        result.update(
            {
                "is_blur": False,
                "variance": 0.0,
                "bubble_count": 0,
                "scratch_count": 0,
                "brightness_mean": 0.0,
                "brightness_std": 0.0,
                "contrast_variance": 0.0,
                "coverage_rate": 0.0,
                "is_low_coverage": True,
                # 异常时的默认值
                "bg_uniformity_std": 0.0,
                "is_bg_non_uniform": True,
                "clump_count": 0,
                "has_clumps": False,
                "defocus_ratio": 0.0,
                "is_partial_defocus": True,
                "media_color_score": 0.0,
                "is_media_abnormal": True,
            }
        )

    return result


# =============================================================================
# 2. 批量处理与统计 (保持不变)
# =============================================================================


def validate_directory(
    directory: Path,
    pattern: str = FILENAME_PATTERN,
    sample_count: Optional[int] = None,
    **kwargs,
) -> List[Dict[str, Any]]:
    """批量处理目录下所有 .tif 文件"""
    # 修改：确保传入的目录如果是相对路径，能正确解析
    if isinstance(directory, str):
        directory = Path(directory)

    if not directory.is_absolute():
        # 如果不是绝对路径，尝试相对于当前工作目录解析，或者保持原样由用户确认
        # 这里主要确保逻辑健壮性，原有逻辑已兼容 Path 对象
        pass

    if not directory.exists():
        raise FileNotFoundError(f"目录不存在：{directory.absolute()}")

    files = sorted(
        [f for f in directory.iterdir() if f.is_file() and f.suffix.lower() == ".tif"]
    )

    if not files:
        print(f"未在 {directory} 中找到任何 .tif 文件")
        return []

    if sample_count and sample_count < len(files):
        random.seed(42)
        files = random.sample(files, sample_count)
        print(f"已启用采样模式，随机抽取 {sample_count} 个文件进行分析")

    results = []
    total = len(files)

    print(f"\n开始处理 {total} 个文件...")
    for i, fpath in enumerate(files):
        print(f"[{i+1}/{total}] 分析：{fpath.name}", end="\r")
        res = analyze_single_image(fpath, pattern, **kwargs)
        results.append(res)

        if not res["readable"]:
            status = "✗ 读取失败"
        elif res["is_qualified"]:
            status = "✓ 合格"
        else:
            issues = []
            if res["is_blur"]:
                issues.append("模糊")
            if res["bubble_count"] > 0:
                issues.append(f"气泡:{res['bubble_count']}")
            if res["scratch_count"] > 0:
                issues.append(f"划痕:{res['scratch_count']}")
            if res["is_low_coverage"]:
                issues.append("低覆盖率")
            status = f"✗ 不合格 ({', '.join(issues)})"
        print(f"[{i+1}/{total}] {fpath.name}: {status}   ")

    return results


def calculate_statistics(results: List[Dict]) -> Dict[str, Any]:
    """计算整体统计指标"""
    total = len(results)
    readable_list = [r for r in results if r.get("readable")]
    readable_count = len(readable_list)

    qualified_count = sum(1 for r in readable_list if r.get("is_qualified"))

    stats = {
        "total_files": total,
        "readable_files": readable_count,
        "unreadable_files": total - readable_count,
        "qualified_files": qualified_count,
        "unqualified_files": readable_count - qualified_count,
        "valid_filename_count": sum(1 for r in results if r.get("filename_valid")),
        "blur_count": sum(1 for r in readable_list if r.get("is_blur")),
        "bubble_count": sum(1 for r in readable_list if r.get("bubble_count", 0) > 0),
        "scratch_count": sum(1 for r in readable_list if r.get("scratch_count", 0) > 0),
        "low_coverage_count": sum(
            1 for r in readable_list if r.get("is_low_coverage", False)
        ),
        # 新增统计项
        "bg_non_uniform_count": sum(
            1 for r in readable_list if r.get("is_bg_non_uniform", False)
        ),
        "clump_detected_count": sum(
            1 for r in readable_list if r.get("has_clumps", False)
        ),
        "partial_defocus_count": sum(
            1 for r in readable_list if r.get("is_partial_defocus", False)
        ),
        "media_abnormal_count": sum(
            1 for r in readable_list if r.get("is_media_abnormal", False)
        ),
        "qualification_rate": (
            (qualified_count / readable_count * 100) if readable_count > 0 else 0.0
        ),
    }

    variances = [r.get("variance", 0) for r in readable_list]
    if variances:
        stats["variance_mean"] = float(np.mean(variances))
        stats["variance_std"] = float(np.std(variances))
        stats["variance_min"] = float(np.min(variances))
        stats["variance_max"] = float(np.max(variances))
    else:
        stats.update(
            {
                "variance_mean": 0,
                "variance_std": 0,
                "variance_min": 0,
                "variance_max": 0,
            }
        )

    # 新增统计：背景均匀性标准差均值
    bg_stds = [r.get("bg_uniformity_std", 0) for r in readable_list]
    if bg_stds:
        stats["bg_std_mean"] = float(np.mean(bg_stds))
    else:
        stats["bg_std_mean"] = 0.0

    # 新增统计：离焦比例均值
    defocus_ratios = [r.get("defocus_ratio", 0) for r in readable_list]
    if defocus_ratios:
        stats["defocus_ratio_mean"] = float(np.mean(defocus_ratios))
    else:
        stats["defocus_ratio_mean"] = 0.0

    return stats


# =============================================================================
# 3. 结果导出 (保持不变)
# =============================================================================


def save_results_to_excel(
    results: List[Dict[str, Any]],
    statistics: Dict[str, Any],
    output_path: Optional[Path] = None,
) -> Path:
    """将分析结果保存为 Excel 文件"""
    if output_path is None:
        DEFAULT_RESULT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = DEFAULT_RESULT_DIR / f"image_validation_{timestamp}.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws_detail = wb.create_sheet("详细检测结果")
    headers = [
        "文件名",
        "文件路径",
        "文件名规范",
        "细胞系",
        "编号",
        "时间",
        "传代",
        "孔位",
        "位置",
        "倍数",
        "可读性",
        "格式",
        "宽",
        "高",
        "是否合格",
        "是否模糊",
        "拉普拉斯方差",
        "气泡数量",
        "划痕数量",
        "覆盖率 (%)",
        "覆盖是否达标",
        "亮度均值",
        "亮度标准差",
        "对比度方差",
        # 新增表头
        "背景均匀性 (Std)",
        "背景是否不均",
        "细胞团块数",
        "是否有团块",
        "离焦区域占比",
        "是否局部离焦",
        "培养基颜色评分",
        "培养基是否异常",
        "错误信息",
    ]
    ws_detail.append(headers)

    for r in results:
        row = [
            r.get("filename"),
            r.get("file_path"),
            "是" if r.get("filename_valid") else "否",
            r.get("cell", ""),
            r.get("id", ""),
            r.get("time", ""),
            r.get("passage", ""),
            r.get("well", ""),
            r.get("location", ""),
            r.get("magnification", ""),
            "是" if r.get("readable") else "否",
            r.get("image_format", ""),
            r.get("width", 0),
            r.get("height", 0),
            "是" if r.get("is_qualified") else "否",
            "是" if r.get("is_blur") else "否",
            f"{r.get('variance', 0):.2f}" if r.get("readable") else "",
            r.get("bubble_count", 0),
            r.get("scratch_count", 0),
            f"{r.get('coverage_rate', 0):.2f}" if r.get("readable") else "",
            "否" if r.get("is_low_coverage") else "是",
            f"{r.get('brightness_mean', 0):.2f}" if r.get("readable") else "",
            f"{r.get('brightness_std', 0):.2f}" if r.get("readable") else "",
            f"{r.get('contrast_variance', 0):.2f}" if r.get("readable") else "",
            f"{r.get('bg_uniformity_std', 0):.2f}" if r.get("readable") else "",
            "是" if r.get("is_bg_non_uniform") else "否",
            r.get("clump_count", 0),
            "是" if r.get("has_clumps") else "否",
            f"{r.get('defocus_ratio', 0):.2f}" if r.get("readable") else "",
            "是" if r.get("is_partial_defocus") else "否",
            f"{r.get('media_color_score', 0):.2f}" if r.get("readable") else "",
            "是" if r.get("is_media_abnormal") else "否",
            r.get("error", ""),
        ]
        ws_detail.append(row)

    ws_summary = wb.create_sheet("统计汇总")
    summary_data = [
        ["总文件数", statistics["total_files"]],
        ["可读文件数", statistics["readable_files"]],
        ["不可读文件数", statistics["unreadable_files"]],
        ["", ""],
        ["合格图片数", statistics["qualified_files"]],
        ["不合格图片数", statistics["unqualified_files"]],
        ["合格率 (%)", f"{statistics['qualification_rate']:.2f}"],
        ["", ""],
        ["文件名规范数", statistics["valid_filename_count"]],
        ["", ""],
        ["质量问题统计", ""],
        ["模糊图片数", statistics["blur_count"]],
        ["含气泡图片数", statistics["bubble_count"]],
        ["含划痕图片数", statistics["scratch_count"]],
        ["低覆盖率图片数", statistics["low_coverage_count"]],
        ["背景不均匀图片数", statistics["bg_non_uniform_count"]],
        ["含细胞团块图片数", statistics["clump_detected_count"]],
        ["局部离焦图片数", statistics["partial_defocus_count"]],
        ["培养基异常图片数", statistics["media_abnormal_count"]],
        ["", ""],
        ["拉普拉斯方差统计", ""],
        ["平均值", f"{statistics['variance_mean']:.2f}"],
        ["标准差", f"{statistics['variance_std']:.2f}"],
        ["最小值", f"{statistics['variance_min']:.2f}"],
        ["最大值", f"{statistics['variance_max']:.2f}"],
        ["", ""],
        ["背景均匀性统计", ""],
        ["平均背景标准差", f"{statistics['bg_std_mean']:.2f}"],
        ["", ""],
        ["离焦检测统计", ""],
        ["平均离焦占比", f"{statistics['defocus_ratio_mean']:.2f}"],
    ]
    for key, val in summary_data:
        ws_summary.append([key, val])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(
                max_len * 1.5 + 4, 50
            )

    wb.save(output_path)
    print(f"\n报告已生成：{output_path}")
    return output_path


# =============================================================================
# 4. 主入口
# =============================================================================


def main(
    target_dir: Path = DEFAULT_RAW_IMAGE_DIR,
    pattern: str = FILENAME_PATTERN,
    output_excel: Optional[Path] = None,
):
    """执行完整验证流程"""
    print("=" * 80)
    print("预处理 - 图像质量自动化验证")
    print("=" * 80)
    print(f"目标目录：{target_dir.absolute()}")
    print(f"命名模板：{pattern}")
    print("=" * 80)

    try:
        results = validate_directory(target_dir, pattern=pattern)

        if not results:
            print("目录下未检测到图像文件")
            return

        stats = calculate_statistics(results)
        print("\n")
        print("=" * 80)
        print("统计摘要")
        print("=" * 80)
        print(f"总数：{stats['total_files']} | 可读：{stats['readable_files']}")
        print(
            f"合格率：{stats['qualification_rate']:.2f}% ({stats['qualified_files']}/{stats['readable_files']})"
        )
        print(
            f"问题分布 -> 模糊:{stats['blur_count']} | 气泡:{stats['bubble_count']} | 划痕:{stats['scratch_count']} | 低覆盖率:{stats['low_coverage_count']}"
        )

        save_results_to_excel(results, stats, output_path=output_excel)

        print("\n验证流程结束。")

    except Exception as e:
        print(f"\n发生严重错误：{e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
